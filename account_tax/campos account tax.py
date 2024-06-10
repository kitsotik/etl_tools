import xmlrpc.client
from openpyxl import Workbook

# Conexión a la instancia de Odoo
url = 'https://56324-16-0-all.runbot.adhoc.com.ar'
db = '56324-16-0-all'
username = 'admin'
password = 'admin'

# Iniciar la conexión XML-RPC
common = xmlrpc.client.ServerProxy(url)
uid = common.authenticate(db, username, password, {})

# Crear una conexión XML-RPC al objeto models de Odoo
models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

# Obtener el modelo account.tax
model_name = 'account.tax'
model_id = models.execute_kw(db, uid, password, model_name, 'search', [[]])[0]

# Obtener los campos del modelo
fields = models.execute_kw(db, uid, password, model_name, 'fields_get', [], {'attributes': ['string', 'type']})

# Crear un archivo de Excel y agregar una hoja
wb = Workbook()
ws = wb.active

# Escribir los nombres de los campos en la primera fila del archivo Excel
for i, field_name in enumerate(fields.keys()):
    ws.cell(row=1, column=i+1, value=field_name)

# Obtener todos los registros del modelo account.tax
records = models.execute_kw(db, uid, password, model_name, 'search_read', [[('id', '>', 0)]], {'fields': list(fields.keys())})

# Escribir los datos de los registros en el archivo Excel
for row_num, record in enumerate(records, start=2):
    for col_num, field_name in enumerate(fields.keys()):
        value = record.get(field_name, '')
        if isinstance(value, list):
            value = value[1] # Extraer el segundo elemento de la lista
        ws.cell(row=row_num, column=col_num+1, value=value)

# Guardar el archivo Excel
wb.save('account_tax.xlsx')
